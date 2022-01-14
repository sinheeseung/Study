import requests
import openpyxl
from bs4 import BeautifulSoup

wb = openpyxl.Workbook()
ws = wb.active  # 현재 활성화된 시트 선택

ws['A1'] = '종목'
ws['B1'] = '현재가'

codes = [
    '005930',
    '000660',
    '035720'
]

companys = [
    '삼성전자',
    'SK하이닉스',
    '카카오'
]

for i in range(len(codes)):
    url = "https://finance.naver.com/item/sise.naver?code=" + codes[i]
    response = requests.get(url)
    html = response.text
    soup = BeautifulSoup(html,'html.parser')
    price = soup.select_one("#_nowVal").text
    price = price.replace(',','')
    ws['A'+str(i+2)] = companys[i]
    ws['B'+str(i+2)] = price

wb.save(r'D:\신희승\Study\Crawling\주식_data.xlsx')