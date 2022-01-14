import openpyxl

# 엑셀 파일 생성
wb = openpyxl.Workbook()

# 엑셀 워크시크 생성(워크시크 이름)
ws = wb.create_sheet('오징어 게임')

# 데이터 추가하기
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = '1'
ws['B2'] = '오일남'

# 엑셀 저장하기(파일이름)
wb.save(r'D:\신희승\Study\Crawling\파이썬 엑셀\참가자_data.xlsx')